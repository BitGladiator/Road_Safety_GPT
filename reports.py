import os
import json
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference
import sqlite3

class ReportGenerator:
    def __init__(self, data_dir='data'):
        self.data_dir = data_dir
        self.analytics_db_path = os.path.join(data_dir, 'analytics.db')
        
    def get_analytics_data(self):
        """Get analytics data from SQLite database"""
        conn = sqlite3.connect(self.analytics_db_path)
        cursor = conn.cursor()
        
        # Get basic stats
        cursor.execute('SELECT COUNT(*) FROM user_queries')
        total_reports = cursor.fetchone()[0]
        
        # Get top interventions
        cursor.execute('''
            SELECT intervention_name, COUNT(*) as count 
            FROM query_interventions 
            GROUP BY intervention_name 
            ORDER BY count DESC 
            LIMIT 10
        ''')
        top_interventions = [{'name': row[0], 'count': row[1]} for row in cursor.fetchall()]
        
        # Get problem types
        cursor.execute('''
            SELECT problem_type, COUNT(*) as count 
            FROM query_interventions 
            GROUP BY problem_type 
            ORDER BY count DESC 
            LIMIT 8
        ''')
        problem_types = [{'name': row[0], 'count': row[1]} for row in cursor.fetchall()]
        
        # Get categories
        cursor.execute('''
            SELECT category, COUNT(*) as count 
            FROM query_interventions 
            GROUP BY category 
            ORDER BY count DESC 
            LIMIT 8
        ''')
        categories = [{'name': row[0], 'count': row[1]} for row in cursor.fetchall()]
        
        conn.close()
        
        return {
            'total_reports': total_reports,
            'top_interventions': top_interventions,
            'problem_types': problem_types,
            'categories': categories,
            'generated_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
    
    def generate_pdf_report(self, output_path=None):
        """Generate PDF compliance report"""
        if not output_path:
            output_path = os.path.join(self.data_dir, f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf')
        
        data = self.get_analytics_data()
        
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.HexColor('#10a37f'),
            alignment=1  # Center
        )
        title = Paragraph("Road Safety Interventions Compliance Report", title_style)
        story.append(title)
        
        # Date
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.gray,
            alignment=1
        )
        date_text = Paragraph(f"Generated on: {data['generated_date']}", date_style)
        story.append(date_text)
        story.append(Spacer(1, 20))
        
        # Summary Stats
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=12,
            textColor=colors.black
        )
        
        story.append(Paragraph(f"<b>Total Reports Generated:</b> {data['total_reports']}", summary_style))
        story.append(Spacer(1, 10))
        
        # Top Interventions Table
        story.append(Paragraph("<b>Most Recommended Interventions</b>", styles['Heading2']))
        intervention_data = [['Intervention', 'Recommendation Count']]
        for intervention in data['top_interventions']:
            intervention_data.append([intervention['name'], str(intervention['count'])])
        
        intervention_table = Table(intervention_data, colWidths=[4*inch, 1.5*inch])
        intervention_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10a37f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(intervention_table)
        story.append(Spacer(1, 20))
        
        # Problem Types
        story.append(Paragraph("<b>Most Common Problem Types</b>", styles['Heading2']))
        problem_data = [['Problem Type', 'Occurrence Count']]
        for problem in data['problem_types']:
            problem_data.append([problem['name'], str(problem['count'])])
        
        problem_table = Table(problem_data, colWidths=[4*inch, 1.5*inch])
        problem_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d8c6c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(problem_table)
        
        doc.build(story)
        return output_path
    
    def generate_excel_report(self, output_path=None):
        """Generate Excel report with analytics and cost estimation"""
        if not output_path:
            output_path = os.path.join(self.data_dir, f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        data = self.get_analytics_data()
        
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = "Road Safety Interventions Analytics Report"
        ws_summary['A1'].font = Font(size=16, bold=True, color='FF10A37F')
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        
        ws_summary['A3'] = "Generated on:"
        ws_summary['B3'] = data['generated_date']
        ws_summary['A4'] = "Total Reports:"
        ws_summary['B4'] = data['total_reports']
        
        # Top Interventions Sheet
        ws_interventions = wb.create_sheet("Top Interventions")
        ws_interventions.append(['Intervention', 'Recommendation Count', 'Priority Level', 'Estimated Cost'])
        
        # Add sample cost estimation and priority
        cost_ranges = {
            'low': '₹5,000 - ₹50,000',
            'medium': '₹50,000 - ₹2,00,000', 
            'high': '₹2,00,000 - ₹10,00,000'
        }
        
        for intervention in data['top_interventions']:
            # Simple priority based on recommendation count
            if intervention['count'] > 5:
                priority = 'High'
                cost = cost_ranges['high']
            elif intervention['count'] > 2:
                priority = 'Medium'
                cost = cost_ranges['medium']
            else:
                priority = 'Low'
                cost = cost_ranges['low']
                
            ws_interventions.append([
                intervention['name'],
                intervention['count'],
                priority,
                cost
            ])
        
        # Style the interventions sheet
        for row in ws_interventions.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='10A37F', end_color='10A37F', fill_type='solid')
        
        # Problem Types Sheet
        ws_problems = wb.create_sheet("Problem Types")
        ws_problems.append(['Problem Type', 'Occurrence Count', 'Severity Level'])
        
        for problem in data['problem_types']:
            # Severity based on occurrence
            if problem['count'] > 4:
                severity = 'Critical'
            elif problem['count'] > 2:
                severity = 'High'
            else:
                severity = 'Medium'
                
            ws_problems.append([problem['name'], problem['count'], severity])
        
        # Style problems sheet
        for row in ws_problems.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='0D8C6C', end_color='0D8C6C', fill_type='solid')
        
        # Auto-adjust column widths
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        return output_path
    
    def generate_compliance_checklist(self, interventions):
        """Generate a compliance checklist for specific interventions"""
        checklist = {
            'title': 'Road Safety Compliance Checklist',
            'generated_date': datetime.now().strftime('%Y-%m-%d'),
            'items': []
        }
        
        for intervention in interventions:
            checklist['items'].append({
                'intervention': intervention['intervention_name'],
                'standard': f"{intervention['standard_code']} Clause {intervention['clause']}",
                'category': intervention['category'],
                'compliance_status': 'Pending Review',
                'priority': self._calculate_priority(intervention),
                'estimated_timeline': '2-4 weeks'
            })
        
        return checklist
    
    def _calculate_priority(self, intervention):
        """Calculate priority based on intervention characteristics"""
        # Simple priority calculation - can be enhanced
        priority_factors = {
            'Traffic Signs': 'High',
            'Road Markings': 'Medium', 
            'Pedestrian Facilities': 'High',
            'Speed Management': 'High',
            'Lighting': 'Medium',
            'Drainage': 'Low'
        }
        
        return priority_factors.get(intervention['category'], 'Medium')

# Example usage
if __name__ == "__main__":
    generator = ReportGenerator()
    
    # Generate sample reports
    pdf_path = generator.generate_pdf_report()
    excel_path = generator.generate_excel_report()
    
    print(f"PDF report generated: {pdf_path}")
    print(f"Excel report generated: {excel_path}")